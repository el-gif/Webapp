import io
import os
os.environ['FOR_DISABLE_CONSOLE_CTRL_HANDLER'] = 'T' # to disable Ctrl+C crashing python when having scipy.interpolate imported (disables Fortrun runtime library from intercepting Ctrl+C signal and lets it to the Python interpreter)
from openpyxl import Workbook
from shiny import ui, App, reactive, render
from shinywidgets import render_widget, output_widget
from ipyleaflet import Map, Marker, MarkerCluster, WidgetControl, FullScreenControl, AwesomeIcon, Heatmap, ScaleControl
from ipywidgets import SelectionSlider, Play, VBox, jslink, Layout, HTML  # pip install ipywidgets==7.6.5, because version 8 has an issue with popups (https://stackoverflow.com/questions/75434737/shiny-for-python-using-add-layer-for-popus-from-ipyleaflet)
import numpy as np
import pandas as pd
import xarray as xr
from scipy.interpolate import CubicSpline, interp2d  # pip install scipy==1.13.1, interp2d much faster than RegularGridInterpolator, even if deprecated
import base64
from ecmwf.opendata import Client
import matplotlib.pyplot as plt
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
import math
import joblib
import torch
import torch.nn as nn
import datetime
import matplotlib.dates as mdates
import sys

# Define initial variables
lat_min, lat_max = 35, 72
lon_min, lon_max = -25, 45


if os.getenv("RENDER") or os.getenv("WEBSITE_HOSTNAME"):  # for Render or Azure Server

    root = "/home/data"

    dir = os.path.join(root, "weather_forecast")

    # Get a list of all .grib files in the directory
    grib_files = [f for f in os.listdir(dir) if f.endswith(".grib")]

    # Count the number of GRIB files
    num_files = len(grib_files)

    # Handle cases where the number of grib files is not exactly 1
    if num_files != 1:
        print(f"Error: {num_files} weather forecasts available. Stopping execution.")
        file = None  # No valid file to use
        sys.exit() # stop execution
    else:
        file = grib_files[0]  # Set 'file' to the only GRIB file
        print(f"Using weather forecast file: {file}")

    new_file_path = os.path.join(dir, file)

else: # if executed locally, data first have to be downloaded

    root = "data"

    # Initialise ECMWF client
    client = Client()
    overwrite = 0  # force overwriting of data
    
    # Determine the latest available forecast based on the ECMWF dissemination schedule https://confluence.ecmwf.int/display/DAC/Dissemination+schedule
    current_utc = datetime.datetime.now(datetime.timezone.utc)

    if current_utc.hour >= 21:
        latest_time = 12  # 12 UTC run is available after 18:27 UTC + 1 hour (+ margin)
        latest_date = current_utc.date()
    elif current_utc.hour >= 9:
        latest_time = 0  # 00 UTC run is available after 06:27 UTC + 1 hour (+ margin)
        latest_date = current_utc.date()
    else:
        latest_time = 12
        latest_date = (current_utc - datetime.timedelta(days=1)).date()

    print(f"Latest available forecast run: {latest_date}, {latest_time} UTC")

    save_dir = os.path.join(root, "weather_forecast")
    os.makedirs(save_dir, exist_ok=True)
    new_file = f"forecast_{latest_date}_{latest_time}.grib"
    new_file_path = os.path.join(save_dir, new_file)

    # Check if the new forecast file is already available
    if os.path.exists(new_file_path) and overwrite == 0:
        print(f"Latest forecast file {new_file} is already available. No download needed.")
    else:
        for old_file in os.listdir(root):
            if old_file.startswith("forecast"):  # Filter files
                old_file_path = os.path.join(root, old_file)  # Get full path
                if os.path.isfile(old_file_path):  # Ensure it's a file (not a folder)
                    print(f"Deleting old file: {old_file}")
                    os.remove(old_file_path)
                    print("File deleted.")


        # Download the new forecast file
        print(f"Downloading new forecast file {new_file}")
        
        # Fetch the latest ECMWF forecast data
        result = client.retrieve(
            type="fc",
            param=["100v", "100u"],  # U- and V-components of wind speed
            target=new_file_path,
            time=latest_time,  # Use the latest available forecast run
            step=list(range(0, 145, 3))
        )

        print(f"New forecast file {new_file} successfully downloaded.")


# Load the wind data (Grib2 file)
ds = xr.open_dataset(new_file_path, engine='cfgrib')

# Filter the data for Europe and extract relevant columns
ds_filtered = ds.sel(latitude=slice(lat_max, lat_min), longitude=slice(lon_min, lon_max))
lats = ds_filtered['latitude'].values
lons = ds_filtered['longitude'].values
u = ds_filtered['u100'].values
v = ds_filtered['v100'].values
valid_times = ds_filtered['valid_time'].values

# Filter the data for Europe and extract relevant columns
df = pd.read_parquet("data/data_hosting/The_Wind_Power.parquet") # 0.7 seconds when WPPs already regionally filtered and stored as parquet file. As unfiltered excel file it takes 11 seconds
df = df.iloc[::100] # only every 10th wpp is possible to alleviate computational and storage burden, not much more
ids = df['ID'].values
countries = df['Country'].values
project_names = df['Name'].values
lats_plants = df['Latitude'].values
lons_plants = df['Longitude'].values
manufacturers = df['Manufacturer'].values
turbine_types = df["Turbine"].replace(["nan", np.nan], "unknown").values
hub_heights = df['Hub height'].values
numbers_of_turbines = df['Number of turbines'].values
capacities = df['Total power'].values / 1e3 # kW to MW as the model has been trained on and the capacity scaler has been fitted
developers = df['Developer'].values
operators = df['Operator'].values
owners = df['Owner'].values
commissioning_dates = df['Commissioning date'].values
ages_months = df['Ages months'].values
commissioning_date_statuses = df['Commissioning date status'].values
hub_height_statuses = df['Hub height status'].values
number_wpps = len(ids)

# Lade die gespeicherte Reihenfolge der Turbinentypen
encoder = joblib.load("modelB/parameters_deployment/encoder.pkl")
known_turbine_types = encoder.categories_[0]
selectable_turbine_types = np.concatenate((known_turbine_types, np.array(["unknown"])))

hub_height_min = math.floor(0.9 * df['Hub height'].min())
hub_height_max = math.ceil(1.1 * df['Hub height'].max())
commissioning_years = df['Commissioning date'].str.split('/').str[0].astype(int)
min_year = commissioning_years.min()
max_year = commissioning_years.max()
min_capacity = 0
max_capacity = capacities.max()

# Lade die Scaler
scalers = joblib.load("modelB/parameters_deployment/scalers.pkl")

# Wende sie auf neue Daten an
scaled_ages_months = scalers["ages"].transform(ages_months.reshape(-1, 1)).flatten()
scaled_hub_heights = scalers["hub_heights"].transform(hub_heights.reshape(-1, 1)).flatten()

class MLP(nn.Module):
    def __init__(self, input_size):
        super(MLP, self).__init__()
        self.fc1 = nn.Linear(input_size, 256)
        self.fc2 = nn.Linear(256, 128)
        self.fc3 = nn.Linear(128, 64)
        self.fc4 = nn.Linear(64, 1)
        self.relu1 = nn.ReLU()
        self.relu2 = nn.ReLU()
        self.relu3 = nn.ReLU()
        self.dropout = nn.Dropout(0.3366)

    def forward(self, x):
        x = self.relu1(self.fc1(x))
        x = self.relu2(self.fc2(x))
        x = self.relu3(self.fc3(x))
        x = self.dropout(x)
        x = self.fc4(x)
        return x
    
# Lade die Metadaten
input_size = joblib.load("modelB/parameters_deployment/input_size.pkl")
model_state_dict = torch.load("modelB/parameters_deployment/model.pth", weights_only=True)
model = MLP(input_size)
model.load_state_dict(model_state_dict)
model.eval()
print("Model loaded")

# Interpolation period and interval
start_time = valid_times[0]
end_time = valid_times[-1]
total_hours = int((end_time - start_time) / np.timedelta64(1, 'h'))
step_size = np.timedelta64(1, 'h')
step_size_hours = step_size / np.timedelta64(1, 'h')
valid_times_interpol = [start_time + i * step_size for i in range(int(total_hours / step_size_hours))]

# Calculate total wind speed and convert to 3D array
total_selection = np.array([np.sqrt(u_value**2 + v_value**2) for u_value, v_value in zip(u, v)])

# Function for temporal interpolation
def interpolation(list):
    list_interpol = []
    for i in range(len(list) - 1):
        step_start = valid_times[i]
        step_end = valid_times[i + 1]
        list_start = list[i]
        list_end = list[i + 1]
        list_interpol.append(list_start)
        interval = step_end - step_start
        for step in np.arange(step_start + step_size, step_end, step_size):
            factor = (step - step_start) / interval
            interpolated_value = (1 - factor) * list_start + factor * list_end
            list_interpol.append(interpolated_value)
    list_interpol.append(list[-1])
    return list_interpol

# Interpolation
total_selection_interpol = interpolation(total_selection)

valid_times_interpol = []
steps = int(total_hours / step_size_hours)
for i in range(steps):
    valid_times_interpol.append(start_time + i * step_size)

example_data = pd.read_parquet("data/data_hosting/example_time_series.parquet")
example_dates = example_data['Date']
example_production = example_data['Production (kW)']

# Read the HTML documentation file
with open("modelB/documentation.html", "r", encoding="utf-8") as f:
    documentation_html = f.read()

app_ui = ui.page_navbar(
    ui.nav_panel(
        "Spatial Forecast",
        output_widget("map")
    ),
    ui.nav_panel(
        "Temporal Forecast",
        ui.row(
            ui.column(2,  # Left column for input fields
                ui.input_slider("lat", "Turbine Latitude", min=lat_min, max=lat_max, value=(lat_min + lat_max) / 2, step=0.01),
                ui.input_slider("lon", "Turbine Longitude", min=lon_min, max=lon_max, value=(lon_min + lon_max) / 2, step=0.01),
                ui.input_select("turbine_type", "Turbine Type", choices=selectable_turbine_types.tolist(), selected=known_turbine_types[0]),
                ui.input_slider("hub_height", "Turbine Hub Height (m)", min=hub_height_min, max=hub_height_max, value=(hub_height_min + hub_height_max) / 2, step=0.1),
                ui.input_slider("commissioning_date_year", "Commissioning Date (Year)", min=min_year, max=max_year, value=(min_year + max_year) / 2, step=1, sep=''),
                ui.input_slider("commissioning_date_month", "Commissioning Date (Month)", min=1, max=12, value=6, step=1),
                ui.input_slider("capacity", "Capacity (MW)", min=min_capacity, max=max_capacity, value=(min_capacity + max_capacity) / 2, step=0.01),
                ui.tags.br(),
                ui.input_file("upload_file", "Contribute data for this configuration", accept=[".xlsx"]),
                ui.input_action_link("download_example", "Download Example File")
            ),
            ui.column(10,  # Right column for output
                ui.panel_well(  # Panel to centre the content
                    ui.output_ui("output_summary"),
                    ui.tags.br(),
                    ui.output_plot("output_graph"),
                    ui.tags.br(),
                    ui.input_action_button("action_button", "Download Forecast")
                ),
            )
        ),
        value='customise_WPP'
    ),
    ui.nav_panel(
        "Documentation",
        ui.HTML(documentation_html)
    ),
    ui.head_content(
        ui.tags.link(rel="icon", type="image/png", href="/WPP_icon2.png"), # image source: https://www.kroschke.com/windsack-set-inkl-korb-und-huelle-korbdurchmesser-650mm-laenge-3500mm--m-8509.html
        ui.tags.script("""
            Shiny.addCustomMessageHandler("download_file", function(message) {
                var link = document.createElement('a');
                link.href = 'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,' + message.data;
                link.download = message.filename;
                document.body.appendChild(link);
                link.click();
                document.body.removeChild(link);
            });
        """)
    ),
    id="navbar_selected",
    title="Wind Power Forecast"
)

# server function
def server(input, output, session):

    ##### Page 1 #####

    # Define reactive values for storing additional information
    project_name = reactive.Value(None)
    operator = reactive.Value(None)
    owner = reactive.Value(None)
    commissioning_date_status = reactive.Value(None)
    hub_height_status = reactive.Value(None)
    country = reactive.Value(None)
    number_turbines = reactive.Value(None)

    is_programmatic_change = reactive.Value(False)

    @reactive.effect
    @reactive.event(input.entire_forecast)
    def entire_forecast_function():

        id = input.entire_forecast()['id']
        
        # Setzen des Flags: Änderungen sind programmatisch
        is_programmatic_change.set(True)

        index = list(ids).index(id)

        # Speichern der zusätzlichen Informationen in den reaktiven Werten
        project_name.set(project_names[index])
        operator.set(operators[index])
        owner.set(owners[index])
        commissioning_date_status.set(commissioning_date_statuses[index])
        hub_height_status.set(hub_height_statuses[index])
        country.set(countries[index])
        number_turbines.set(numbers_of_turbines[index])

        # Parameter extrahieren
        lat = lats_plants[index]
        lon = lons_plants[index]
        turbine_type = turbine_types[index]
        hub_height = hub_heights[index]
        capacity = capacities[index]
        commissioning_date = commissioning_dates[index]
        commissioning_date_year, commissioning_date_month = commissioning_date.split("/")
        commissioning_date_year = int(commissioning_date_year)
        commissioning_date_month = int(commissioning_date_month)

        # Update der Eingabefelder mit neuen Werten vor Wechseln des Tabs.
        # Attention: the rounding must exactly correspond to the steps, the sliders have been initialised with.
        # Otherwise, a rounding will occur automatically, and the function observe_slider_changes() is called an additional time,
        # while is_programmatic_change is already false --> output_summary will be reset --> to avoid
        ui.update_slider("lat", value=round(lat, 2))
        ui.update_slider("lon", value=round(lon, 2))
        ui.update_select("turbine_type", selected=turbine_type if turbine_type in selectable_turbine_types else "unknown")
        ui.update_slider("hub_height", value=round(hub_height, 1))
        ui.update_slider("commissioning_date_year", value=commissioning_date_year)
        ui.update_slider("commissioning_date_month", value=commissioning_date_month)
        ui.update_slider("capacity", value=(capacity, 2))

        # Wechsel zu "Customise WPP" Tab
        ui.update_navs("navbar_selected", selected="customise_WPP")

    @render_widget
    def map():
        m = Map(
            center=[(lat_min + lat_max) / 2, (lon_min + lon_max) / 2],
            zoom=5,
            layout=Layout(width='100%', height='90vh'),
            scroll_wheel_zoom=True
        )

        scale = ScaleControl(position="bottomleft", imperial=False)
        m.add(scale)

        markers = []
        for name, capacity, number_of_turbines, turbine_type, operator, id, lat, lon in zip(project_names, capacities, numbers_of_turbines, turbine_types, operators, ids, lats_plants, lons_plants):
            
            popup_content = HTML(
                f"<strong>Project Name:</strong> {name}<br>"
                f"<strong>Capacity:</strong> {capacity} MW<br>"
                f"<strong>Number of turbines:</strong> {number_of_turbines}<br>"
                f"<strong>Turbine Type:</strong> {turbine_type}<br>"
                f"<strong>Operator:</strong> {operator}<br>"
                f"<strong>Wind speed forecast:</strong> select forecast step<br>"
                f"<strong>Production forecast:</strong> select forecast step<br>"
                f"<button onclick=\"Shiny.setInputValue('entire_forecast', {{id: {id}, timestamp: Date.now()}})\">Entire Forecast</button>" # timestamp to always have a slightly different button value to ensure that each and every click on the "entire forecast" button triggers the event, even click on same button twice
            )

            marker = Marker(
                location=(lat, lon),
                popup=popup_content,
                rise_offset=True,
                draggable=False
            )
            markers.append(marker)

        # Cluster erstellen und zur Karte hinzufügen
        marker_cluster = MarkerCluster(markers=markers)
        m.add(marker_cluster)

        # Slider for time steps
        play = Play(min=0, max=total_hours, step=step_size_hours, value=0, interval=500, description='Time Step')
        slider = SelectionSlider(options=valid_times_interpol, value=valid_times_interpol[0], description='Time')
        jslink((play, 'value'), (slider, 'index'))
        slider_box = VBox([play, slider])
        m.add(WidgetControl(widget=slider_box, position='topright'))

        # Map update function based on slider
        def update_map(change):
            time_step = slider.value
            step_index = int((time_step - start_time) / step_size)
            spatial_interpolator = interp2d(lons, lats, total_selection_interpol[step_index], kind='cubic')
            wind_speeds_at_points = np.array([spatial_interpolator(lon, lat)[0] for lon, lat in zip(lons_plants, lats_plants)])
            scaled_wind_speeds_at_points = scalers["winds"].transform(wind_speeds_at_points.reshape(-1, 1)).flatten()

            turbine_types_onehot = np.zeros((number_wpps, len(known_turbine_types)))
            for i, turbine_type in enumerate(turbine_types):
                if turbine_type not in known_turbine_types:
                    turbine_types_onehot[i] = np.full(len(known_turbine_types), 1.0 / len(known_turbine_types))
                else:
                    turbine_types_onehot[i] = encoder.transform(np.array([[turbine_type]])).flatten()

            all_input_features = np.hstack([
                turbine_types_onehot,
                scaled_hub_heights.reshape(-1, 1),
                scaled_ages_months.reshape(-1, 1),
                scaled_wind_speeds_at_points.reshape(-1, 1)
            ])

            input_tensor = torch.tensor(all_input_features, dtype=torch.float32)

            with torch.no_grad():
                predictions = np.minimum(model(input_tensor).numpy().flatten() * capacities, capacities)

            predictions[(predictions < 0)] = 0
            
            # Update marker pop-ups with production values
            for marker, name, capacity, number_of_turbines, turbine_type, operator, wind_speed, prediction, id in zip(marker_cluster.markers, project_names, capacities, numbers_of_turbines, turbine_types, operators, wind_speeds_at_points, predictions, ids):
                marker.popup.value = \
                    f"<strong>Project Name:</strong> {name}<br>"\
                    f"<strong>Capacity:</strong> {capacity} MW<br>"\
                    f"<strong>Number of Turbines:</strong> {number_of_turbines}<br>"\
                    f"<strong>Turbine Type:</strong> {turbine_type}<br>"\
                    f"<strong>Operator:</strong> {operator}<br>"\
                    f"<strong>Wind speed forecast:</strong> {wind_speed:.2f} m/s<br>"\
                    f"<strong>Production forecast:</strong> {prediction:.2f} MW<br>"\
                    f"<button onclick=\"Shiny.setInputValue('entire_forecast', {{id: {id}, timestamp: Date.now()}})\">Entire Forecast</button>" # timestamp to always have a slightly different button value to ensure that each and every click on the "entire forecast" button triggers the event, even click on same button twice
                
            # Prepare heatmap data for "operating" wind plants
            heatmap_data = [(lat, lon, prod) for lat, lon, prod in zip(lats_plants, lons_plants, predictions)]

            # Remove existing heatmap layer if any and add new one
            for layer in m.layers:
                if isinstance(layer, Heatmap):
                    m.remove(layer)
            heatmap = Heatmap(locations=heatmap_data, radius=10, blur=10, max_zoom=10)
            m.add(heatmap)
        
        update_map(None)
        
        # Observe slider value changes and update map
        slider.observe(update_map, names='value')

        # Add FullScreenControl to map
        m.add(FullScreenControl())

        return m
    

    ##### Page 2 #####

    forecast_data = reactive.Value({"wind_speeds": None, "productions": None})
    time_series = reactive.Value(None)
    button_status = reactive.Value("download")

    # Function to handle file upload
    @reactive.effect
    @reactive.event(input.upload_file)
    def handle_file_upload():
        if input.upload_file() is not None:
            try:
                file = input.upload_file()[0]['datapath']

                # Check if the file is empty
                if os.path.getsize(file) == 0:
                    ui.notification_show("File is empty. Please check the file and try again.", duration=None)
                    return

                # Attempt to read Excel file with specified columns
                time_series_data = pd.read_excel(file)

                # Check if DataFrame is empty
                if time_series_data.empty:
                    ui.notification_show("The Excel file contains no data. Please provide a valid time series.", duration=None)
                    return

                required_columns = ["Date", "Production (MW)"]
                if not all(col in time_series_data.columns for col in required_columns):
                    ui.notification_show(f"The Excel file is missing required columns. Please ensure it includes {required_columns}.", duration=None)
                    return

                time_series.set(time_series_data)  # Calls output_graph function
                ui.update_action_button("action_button", label="Contribute Data")
                button_status.set('contribute')

            except FileNotFoundError:
                ui.notification_show("File not found. Please upload a valid file.", duration=None)

            except ValueError as ve:
                ui.notification_show(f"Value error: {str(ve)}. Ensure the file format is correct.", duration=None)

            except pd.errors.ExcelFileError:
                ui.notification_show("The file format is not recognized as an Excel file. Please upload a valid .xlsx or .xls file.", duration=None)

            except Exception as e:
                # Generic catch-all for any other errors
                ui.notification_show(f"An unexpected error occurred: {str(e)}. Please try again.", duration=None)

    
    # Generate example file for download
    @reactive.effect
    @reactive.event(input.download_example)
    async def generate_example_file():
        # Create example workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Example Time Series"
        
        # Add header and example data
        ws.append(["Date", "Production (kW)"])
        for date, production in zip(example_dates, example_production):
            ws.append([date, production])
        
        # Save workbook to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Encode and send for download
        file_data_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        await session.send_custom_message("download_file", {
            "data": file_data_base64,
            "filename": "example_time_series.xlsx"
        })

    # Observing slider changes to revert to forecast view
    @reactive.effect
    @reactive.event(input.lat, input.lon, input.turbine_type, input.hub_height, input.commissioning_date_year, input.commissioning_date_month, input.capacity)
    def observe_slider_changes():
        # Überspringen, wenn Änderungen programmatisch sind
        if is_programmatic_change.get():
            is_programmatic_change.set(None)
            return

        # reset reactive variables
        project_name.set(None)
        operator.set(None)
        owner.set(None)
        commissioning_date_status.set(None)
        hub_height_status.set(None)
        country.set(None)
        number_turbines.set(None)

        # display regular elements
        time_series.set(None) # calls output_graph function
        ui.update_action_button("action_button", label="Download Forecast")
        button_status.set('download')

    # Capture user input and display configuration summary
    @render.text
    def output_summary():
        # Capture inputs
        lat_plant = input.lat()
        lon_plant = input.lon()
        turbine_type = input.turbine_type()
        hub_height = input.hub_height()
        commissioning_date_year = input.commissioning_date_year()
        commissioning_date_month = input.commissioning_date_month()
        capacity = input.capacity()

        if hub_height_status.get() == 1:
            hub_height_display = "nan"
        else: # hub_height_status.get() == 0 or None (custom configuration)
            hub_height_display = f"{hub_height:.2f} m"

        if commissioning_date_status.get() == 2:
            commissioning_date_display = "nan"
        elif commissioning_date_status.get() == 1:
            commissioning_date_display = str(commissioning_date_year)
        else: # commissioning_date_status.get() == 0 or None (custom configuration)
            commissioning_date_display = f"{commissioning_date_year}/{commissioning_date_month}"

        # Return configuration summary text
        summary_html = (
            f"<b>Wind Power Plant Configuration</b><br><br>"
            f"<b>Project Name:</b> {project_name.get()}<br>"
            f"<b>Country:</b> {country.get()}<br>"
            f"<b>Capacity:</b> {capacity} MW<br>"
            f"<b>Number of Turbines:</b> {number_turbines.get()}<br>"
            f"<b>Operator:</b> {operator.get()}<br>"
            f"<b>Owner:</b> {owner.get()}<br>"
            f"<b>Location:</b> ({lat_plant:.2f}, {lon_plant:.2f})<br>"
            f"<b>Type:</b> {turbine_type}<br>"
            f"<b>Hub Height:</b> {hub_height_display}<br>"
            f"<b>Commissioning Date:</b> {commissioning_date_display}<br>"

        )
        return ui.HTML(summary_html)

    # Plot forecasted production over time
    @render.plot
    def output_graph():
        fig, ax1 = plt.subplots(figsize=(8, 4))
        ax1.set_xlabel("Date")
        ax1.set_ylabel("Production (MW)")

        # Add horizontal dashed line for capacity
        capacity = input.capacity()
        ax1.axhline(y=capacity, color='gray', linestyle='--', label=f"Capacity ({capacity} MW)")

        time_series_data = time_series.get()

        if time_series_data is not None and not time_series_data.empty: # check if user has uploaded time series to plot it
            ax1.plot(pd.to_datetime(time_series_data.iloc[:, 0], errors='coerce'), time_series_data.iloc[:, 1] / 1e3, label="Uploaded Time Series (MW)", color='black')
            ax1.set_title("Historical Wind Power Plant Production")

            ax2 = ax1.twinx()
            y_min, y_max = ax1.get_ylim()
            ax2.set_ylim(y_min / capacity, y_max / capacity)
            ax2.set_ylabel("Capacity Factor")
        else: # Retrieve inputs
            lat_plant = input.lat()
            lon_plant = input.lon()
            turbine_type = input.turbine_type()
            hub_height = input.hub_height()
            commissioning_date_year = input.commissioning_date_year()
            commissioning_date_month = input.commissioning_date_month()
            ref_date = pd.Timestamp("2024-12-01")
            age_months = (ref_date.year - commissioning_date_year) * 12 + (ref_date.month - commissioning_date_month)
            capacity = input.capacity()

            scaled_hub_height = scalers["hub_heights"].transform(np.array([[hub_height]]))[0][0]
            scaled_age_months = scalers["ages"].transform(np.array([[age_months]]))[0][0]

            # Calculate forecasted productions
            wind_speeds_at_point = []
            num_steps = len(valid_times_interpol)
            for step_index in range(num_steps):
                spatial_interpolator = interp2d(lons, lats, total_selection_interpol[step_index], kind='cubic')
                wind_speeds_at_point.append(spatial_interpolator(lon_plant, lat_plant)[0])

            scaled_wind_speeds_at_point = scalers["winds"].transform(np.array(wind_speeds_at_point).reshape(-1, 1)).flatten()

            if turbine_type == "unknown":
                num_categories = len(known_turbine_types)
                one_hot_vector = np.full((1, num_categories), 1 / num_categories)  # mixture of all known turbine types
                turbine_type_repeated = np.tile(one_hot_vector, (num_steps, 1))
            else:
                turbine_type_repeated = np.tile(encoder.transform(np.array([[turbine_type]])), (num_steps, 1))

            hub_height_repeated = np.full((num_steps, 1), scaled_hub_height)
            age_repeated = np.full((num_steps, 1), scaled_age_months)
            wind_speed_column = scaled_wind_speeds_at_point.reshape(-1, 1)

            all_input_features = np.hstack([
                turbine_type_repeated,
                hub_height_repeated,
                age_repeated,
                wind_speed_column
            ])

            input_tensor = torch.tensor(all_input_features, dtype=torch.float32)

            with torch.no_grad():
                predictions_cap_factor = np.minimum(model(input_tensor).numpy().flatten(), np.full(num_steps, 1))

            predictions_cap_factor[(predictions_cap_factor < 0)] = 0
            predictions_power = predictions_cap_factor * capacity


            # Create cubic spline interpolation
            cs = CubicSpline(valid_times_interpol, predictions_power)

            # Generate a finer time grid for smooth plotting
            fine_time_grid = pd.date_range(start=valid_times_interpol[0], end=valid_times_interpol[-1], periods=500)

            # Get smooth interpolated values
            smoothed_predictions = np.maximum(cs(fine_time_grid), 0)

            # Store the forecast data in the reactive `forecast_data`
            forecast_data.set({"wind_speeds": wind_speeds_at_point, "productions": smoothed_predictions})

            # Plot original and smoothed data
            #ax1.plot(valid_times_interpol, predictions_power, 'o', label="Original Points", color="red")
            ax1.plot(fine_time_grid, smoothed_predictions, '-', label="Cubic Spline Interpolation", color="blue")
            ax1.set_xlabel("Date")
            ax1.set_ylabel("Production (MW)")
            ax1.set_ylim(bottom=0)
            ax1.set_title("Forecasted Wind Power Plant Production")

            # Add a red dotted vertical line at the current time
            current_time = pd.Timestamp.now()  # or use a specific time if needed
            ax1.axvline(x=current_time, color='red', linestyle='--', linewidth=1.5, label='Current Time')

            ax2 = ax1.twinx()
            y_min, y_max = ax1.get_ylim()
            ax2.set_ylim(y_min / capacity, y_max / capacity)
            ax2.set_ylabel("Capacity Factor")

        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y/%m/%d'))

        # Rotate the labels for better fit
        plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')

        ax1.grid(True)
        plt.tight_layout()
        return fig  # Returning the figure will embed it in the app

    def download_forecast_data():

        data = forecast_data.get()
        wind_speeds, productions = data["wind_speeds"], data["productions"]

        lat_plant = input.lat()
        lon_plant = input.lon()
        turbine_type = input.turbine_type()
        hub_height = input.hub_height()
        commissioning_date_year = input.commissioning_date_year()
        commissioning_date_month = input.commissioning_date_month()
        capacity = input.capacity()

        # Create a new workbook and add data
        wb = Workbook()
        
        # Sheet1: Turbine Specifications
        ws_specs = wb.active
        ws_specs.title = "Turbine Specifications"

        # Define the specs_data list with required values only
        specs_data = [
            ["Specification", "Value"],
            ["Project Name", project_name.get()],
            ["Number of Turbines", number_turbines.get()],
            ["Operator", operator.get()],
            ["Owner", owner.get()],
            ["Location", f"({lat_plant}, {lon_plant})"],
            ["Type", turbine_type],
            ["Hub Height (m)", round(hub_height, 2) if hub_height_status.get() == 0 else "nan"],
            ["Commissioning Date", f"{commissioning_date_year}/{commissioning_date_month}" if commissioning_date_status.get() == 0 else (str(commissioning_date_year) if commissioning_date_status.get() == 1 else "nan")],
            ["Capacity (MW)", capacity]
        ]

        # Populate the workbook
        for row in specs_data:
            ws_specs.append(row)

        # Sheet2: Production Forecast
        ws_forecast = wb.create_sheet("Production Forecast")

        # Add headers for date, wind speed, and production
        ws_forecast.append(["Date", "Wind Speed (m/s)", "Production (MW)"])

        # Add production data per time step
        for time, wind_speed, production in zip(valid_times_interpol, wind_speeds, productions):
            time_as_datetime = pd.to_datetime(time).to_pydatetime()
            ws_forecast.append([time_as_datetime, wind_speed, production])  # Ensure wind speed and production values are correctly added

        # Save in buffer and return
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # Function to trigger download on button click
    @reactive.effect
    @reactive.event(input.action_button)
    async def action():
        if button_status.get() == "download":
        
            # Retrieve forecast data as bytes
            file_data = download_forecast_data()
            
            # Encode file in Base64
            file_data_base64 = base64.b64encode(file_data).decode('utf-8')
            
            # Send file as Base64 to client for download
            await session.send_custom_message("download_file", {
                "data": file_data_base64,
                "filename": "forecasted_production.xlsx"
            })
        
        else: # button_status.get() == "contribute"
            time_series_data = time_series.get()

            # Notify the user that the file is being saved
            ui.notification_show(
                "Thank you for uploading your time series. It will be checked for plausibility and possibly used to improve the model at the next training.",
                duration=None
            )

            # Metadata from input sliders
            metadata = {
                "Latitude": input.lat(),
                "Longitude": input.lon(),
                "Turbine Type": input.turbine_type(),
                "Hub Height": input.hub_height(),
                "Commissioning Year": input.commissioning_date_year(),
                "Commissioning Month": input.commissioning_date_month(),
                "Capacity (MW)": input.capacity(),
            }

            # Save folder and timestamp
            save_dir = os.path.join(root, "crowdsourced_data")
            os.makedirs(save_dir, exist_ok=True)
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            save_path = os.path.join(save_dir, f"time_series_{timestamp}.xlsx")

            # Combine metadata and time series data into a single DataFrame
            metadata_df = pd.DataFrame([metadata])  # Convert metadata to DataFrame

            # Save metadata and time series data to the same Excel file
            with pd.ExcelWriter(save_path, engine='xlsxwriter') as writer:
                metadata_df.to_excel(writer, sheet_name="Metadata", index=False)
                time_series_data.to_excel(writer, sheet_name="Time Series", index=False)

# Define absolute path to `www` folder
path_www = os.path.join(os.path.dirname(__file__), "..", "data/www")

# Start the app with the `www` directory as static content
app = App(app_ui, server, static_assets={"/": path_www})

if __name__ == "__main__":
    # Check if the variable `RENDER` is set to detect if the app is running on Render
    if os.getenv("RENDER") or os.getenv("WEBSITE_HOSTNAME"):  # for Render or Azure Server
        host = "0.0.0.0"  # For Render or other external deployments
    else:
        host = "127.0.0.1"  # For local development (localhost)

    app.run(host=host, port=8000)  # port binding: set the server port to 8000, because this is what Azure expects
