import io
import os
os.environ['FOR_DISABLE_CONSOLE_CTRL_HANDLER'] = 'T' # to disable Ctrl+C crashing python when having scipy.interpolate imported (disables Fortrun runtime library from intercepting Ctrl+C signal and lets it to the Python interpreter)
from openpyxl import Workbook
from shiny import ui, App, reactive, render
from shinywidgets import render_widget, output_widget
from ipyleaflet import Map, Marker, MarkerCluster, WidgetControl, FullScreenControl, AwesomeIcon, Heatmap
from ipywidgets import SelectionSlider, Play, VBox, jslink, Layout, HTML  # pip install ipywidgets==7.6.5, because version 8 has an issue with popups (https://stackoverflow.com/questions/75434737/shiny-for-python-using-add-layer-for-popus-from-ipyleaflet)
import numpy as np
import pandas as pd
import xarray as xr
from scipy.interpolate import interp1d
from scipy.interpolate import interp2d  # pip install scipy==1.13.1, interp2d much faster than RegularGridInterpolator, even if deprecated
import base64
from ecmwf.opendata import Client
from branca.colormap import linear
import matplotlib.pyplot as plt
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
import time
starttime = time.time()

# temporarily
turbine_type = 'Type A'
hub_height = 100
rotor_diameter = 50

# Define initial variables
initial = 0
lat_min, lat_max = 35, 72
lon_min, lon_max = -25, 45

# Path to the stored weather data
wind_file = "data/weather_forecast/data_europe.grib2"
# Set time interval for updates (in seconds)
time_threshold = 24 * 3600  # 24 hours in seconds
overwrite = 0  # force overwriting of data

# Initialise ECMWF client
client = Client(
    source="ecmwf",
    model="ifs",
    resol="0p25"
)

if os.getenv("RENDER") or os.getenv("WEBSITE_HOSTNAME"):  # for Render or Azure Server
    # for saving storage on server, only 4 time steps
    step_selection = [0, 3, 6, 9]
else:
    # all available steps up to 7 days (168 hours)
    step_selection = [0, 3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33, 36, 39, 42, 45, 48, 51, 54, 57, 60,
                      63, 66, 69, 72, 75, 78, 81, 84, 87, 90, 93, 96, 99, 102, 105, 108, 111, 114, 117, 120,
                      123, 126, 129, 132, 135, 138, 141, 144, 150, 156, 162, 168]

# all available steps
# step_selection = [
#     0, 3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33, 36, 39, 42, 45, 48, 51, 54, 57, 60,
#     63, 66, 69, 72, 75, 78, 81, 84, 87, 90, 93, 96, 99, 102, 105, 108, 111, 114, 117, 120,
#     123, 126, 129, 132, 135, 138, 141, 144, 150, 156, 162, 168, 174, 180, 186, 192, 198, 204,
#     210, 216, 222, 228, 234, 240
# ]

# # Function to check the age of the file, to avoid a cron-job, which is a paid feature on most Cloud Computing platforms
# def is_data_stale(wind_file, time_threshold):
#     if not os.path.exists(wind_file):
#         # File does not exist, data must be fetched
#         return True
#     # Determine the age of the file (in seconds since the last modification)
#     file_age = time.time() - os.path.getmtime(wind_file)
#     # Check if the file is older than the specified threshold
#     return file_age > time_threshold

# if is_data_stale(wind_file, time_threshold) or overwrite:
#     print("Data is outdated, does not exist, or should be overwritten. Fetching new data.")
#     # Fetching API data since they are either missing or older than 10 hours
#     result = client.retrieve(  # retrieve data worldwide, because no area tag available (https://github.com/ecmwf/ecmwf-opendata, https://www.ecmwf.int/en/forecasts/datasets/open-data)
#         type="fc",
#         param=["100v", "100u"],  # U- and V-components of wind speed
#         target=wind_file,
#         time=0,  # Forecast time (model run at 00z)
#         step=step_selection
#     )
#     print("New data has been successfully fetched and saved.")
# else:
#     print("Data is current and will be used.")

# Load the wind data (Grib2 file)
ds = xr.open_dataset(wind_file, engine='cfgrib')

# Filter the data for Europe and extract relevant columns
ds_filtered = ds.sel(latitude=slice(lat_max, lat_min), longitude=slice(lon_min, lon_max))
lats = ds_filtered['latitude'].values
lons = ds_filtered['longitude'].values
u = ds_filtered['u100'].values
v = ds_filtered['v100'].values
valid_times = ds_filtered['valid_time'].values

# Filter the data for Europe and extract relevant columns
df_filtered = pd.read_parquet("data/WPPs/Global-Wind-Power-Tracker-Europe.parquet") # 0.7 seconds when WPPs already regionally filtered and stored as parquet file. As unfiltered excel file it takes 11 seconds
df_filtered = df_filtered[df_filtered['Status'] == 'operating'].iloc[::100] # only operating WPPs, and only every 100th to alleviate computational and storage burden
df_filtered['ID'] = list(range(1, len(df_filtered) + 1))
ids = df_filtered['ID'].values
lats_plants = df_filtered['Latitude'].values
lons_plants = df_filtered['Longitude'].values
capacities = df_filtered['Capacity (MW)'].values
project_names = df_filtered['Project Name'].values
statuses = df_filtered['Status'].values
operators = df_filtered['Operator'].values
owners = df_filtered['Owner'].values
commission_dates = df_filtered['Start year'].fillna(2020).values

# Interpolation period and interval
start_time = valid_times[0]
end_time = valid_times[-1]
total_hours = int((end_time - start_time) / np.timedelta64(1, 'h'))
step_size = np.timedelta64(1, 'h')
step_size_hours = step_size / np.timedelta64(1, 'h')
valid_times_interpol = [start_time + i * step_size for i in range(int(total_hours / step_size_hours))]

# Calculate total wind speed and convert to 3D array
total_selection = np.array([np.sqrt(u_value**2 + v_value**2) for u_value, v_value in zip(u, v)])

# Define the power curve (wind speed and output)
wind_speeds = np.arange(0, 25.5, 0.5)
power_output = [0]*7 + [35, 80, 155, 238, 350, 474, 630, 802, 1018, 1234, 1504, 1773, 2076, 2379, 2664, 2948, 3141, 3334, 3425, 3515, 3546, 3577, 3586, 3594, 3598, 3599] + [3600]*18
max_capacity = 3600
power_output_normalised = [value / max_capacity for value in power_output]
power_curve_normalised = interp1d(wind_speeds, power_output_normalised, kind='cubic', fill_value="extrapolate")

# Function for temporal interpolation
def interpolation(list):
    list_interpol = []
    for i in range(len(list) - 1):
        step_start = valid_times[i]
        step_end = valid_times[i + 1]
        list_start = list[i]
        list_end = list[i + 1]
        list_interpol.append(list_start)
        interval = step_end - step_start
        for step in np.arange(step_start + step_size, step_end, step_size):
            factor = (step - step_start) / interval
            interpolated_value = (1 - factor) * list_start + factor * list_end
            list_interpol.append(interpolated_value)
    list_interpol.append(list[-1])
    return list_interpol

# Interpolation
total_selection_interpol = interpolation(total_selection)

valid_times_interpol = []
steps = int(total_hours / step_size_hours)
for i in range(steps):
    valid_times_interpol.append(start_time + i * step_size)

example_data = pd.read_parquet("data/production_history/Example/example_time_series.parquet")
example_dates = example_data['Date']
example_production = example_data['Production (kW)']

app_ui = ui.page_navbar(
    ui.nav_panel(
        "WPP database",
        output_widget("map")
    ),
    ui.nav_panel(
        "Customise WPP",
        ui.row(
            ui.column(2,  # Left column for input fields
                ui.input_numeric("lat", "Turbine Latitude", value=50.0, min=lat_min, max=lat_max),
                ui.input_numeric("lon", "Turbine Longitude", value=10.0, min=lon_min, max=lon_max),
                ui.input_select("turbine_type", "Turbine Type", choices=["Type A", "Type B", "Type C"]),
                ui.input_slider("hub_height", "Turbine Hub Height (m)", min=0, max=200, value=100),
                ui.input_slider("commission_date", "Commission Date", min=1990, max=2024, value=2022, step=1),
                ui.input_slider("rotor_diameter", "Rotor Diameter (m)", min=0, max=100, value=50),
                ui.input_slider("capacity", "Capacity (kW)", min=0, max=20000, value=5000),
                ui.tags.br(),
                ui.input_file("upload_file", "Contribute data for this configuration", accept=[".xlsx"]),
                ui.input_action_link("download_example", "Download Example File")
            ),
            ui.column(10,  # Right column for output
                ui.panel_well(  # Panel zur Zentrierung des Inhalts
                    ui.output_ui("output_summary"),
                    ui.tags.br(),
                    ui.output_plot("output_graph"),
                    ui.tags.br(),
                    ui.input_action_button("action_button", "Download Forecast")
                ),
            )
        ),
        value='customise_WPP'
    ),
    ui.nav_panel(
        "Documentation",
        ui.h3("Wind Power Forecast Application"),
        ui.p(
            "This web application visualises wind power production forecasts for wind power plants across Europe. "
            "Users can view wind plant information, forecasted wind speeds, and production values on an interactive map. "
            "Each wind plant can be customised by selecting its attributes, and predictions are displayed based on input parameters."
        )
    ),
    ui.head_content(
        ui.tags.script("""
            Shiny.addCustomMessageHandler("download_file", function(message) {
                var link = document.createElement('a');
                link.href = 'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,' + message.data;
                link.download = message.filename;
                document.body.appendChild(link);
                link.click();
                document.body.removeChild(link);
            });
        """),
        ui.tags.link(rel="icon", type="image/png", href="/www/WPP_icon2.png")
    ),
    id="navbar_selected",
    title="Wind Power Forecast"
)

# server function
def server(input, output, session):

    ##### Page 1 #####

    # Define reactive values for storing additional information
    project_name = reactive.Value(None)
    status = reactive.Value(None)
    operator = reactive.Value(None)
    owner = reactive.Value(None)

    @reactive.effect
    @reactive.event(input.entire_forecast)
    def entire_forecast_function():
        id = input.entire_forecast()['id']

        index = list(ids).index(id)

        # Speichern der zusätzlichen Informationen in den reaktiven Werten
        project_name.set(project_names[index])
        status.set(statuses[index])
        operator.set(operators[index])
        owner.set(owners[index])

        # Dann die anderen Parameter extrahieren
        lat = lats_plants[index]
        lon = lons_plants[index]
        capacity = capacities[index]
        commission_date = commission_dates[index]

        # Update der Eingabefelder mit neuen Werten vor Wechseln des Tabs
        ui.update_numeric("lat", value=lat)
        ui.update_numeric("lon", value=lon)
        ui.update_select("turbine_type", selected=turbine_type)
        ui.update_slider("hub_height", value=hub_height)
        ui.update_slider("commission_date", value=commission_date)
        ui.update_slider("rotor_diameter", value=rotor_diameter)
        ui.update_slider("capacity", value=capacity)

        # Wechsel zu "Customise WPP" Tab
        ui.update_navs("navbar_selected", selected="customise_WPP")

    @output
    @render_widget
    def map():
        m = Map(
            center=[(lat_min + lat_max) / 2, (lon_min + lon_max) / 2],
            zoom=5,
            layout=Layout(width='100%', height='95vh'),
            scroll_wheel_zoom=True
        )

        markers = []
        for lat, lon, name, capacity, status, operator, owner, commission_date, id in zip(lats_plants, lons_plants, project_names, capacities, statuses, operators, owners, commission_dates, ids):

            # if status == "operating":
            #     color = "green"
            # elif status in ["construction", "pre-construction", "announced"]:
            #     color = "orange"
            # else:
            #     color = "red"

            # icon = AwesomeIcon(
            #     name="circle",
            #     marker_color=color
            # )
            
            popup_content = HTML(
                f"<strong>Project Name:</strong> {name}<br>"
                f"<strong>Status:</strong> {status}<br>"
                f"<strong>Capacity:</strong> {capacity} MW<br>"
                f"<strong>Operator:</strong> {operator}<br>"
                f"<strong>Owner:</strong> {owner}<br>"
                f"<strong>Type:</strong> {turbine_type}<br>"
                f"<strong>Height:</strong> {hub_height} MW<br>"
                f"<strong>Comission Date:</strong> {commission_date}<br>"
                f"<strong>Diameter:</strong> {rotor_diameter}<br>"                    
                f"<strong>Wind speed forecast:</strong> select forecast step<br>"
                f"<strong>Production forecast:</strong> select forecast step<br>"
                f"<button onclick=\"Shiny.setInputValue(\'entire_forecast\', {{id: {id}}})\">Entire Forecast</button>"
            )

            marker = Marker(
                location=(lat, lon),
                popup=popup_content,
                #icon=icon, # takes too long
                rise_offset=True,
                draggable=False
            )
            markers.append(marker)

        # Cluster erstellen und zur Karte hinzufügen
        marker_cluster = MarkerCluster(markers=markers)
        m.add(marker_cluster)

        # Slider for time steps
        play = Play(min=0, max=total_hours, step=step_size_hours, value=0, interval=500, description='Time Step')
        slider = SelectionSlider(options=valid_times_interpol, value=valid_times_interpol[0], description='Time')
        jslink((play, 'value'), (slider, 'index'))
        slider_box = VBox([play, slider])
        m.add(WidgetControl(widget=slider_box, position='topright'))

        # Map update function based on slider
        def update_map(change):
            time_step = slider.value
            step_index = int((time_step - start_time) / step_size)
            print(lons.shape)
            print(lats.shape)
            print(total_selection_interpol[step_index].shape)
            spatial_interpolator = interp2d(lons, lats, total_selection_interpol[step_index], kind='cubic')
            wind_speeds_at_points = [spatial_interpolator(lon, lat)[0] for lon, lat in zip(lons_plants, lats_plants)]

            productions = np.abs(power_curve_normalised(wind_speeds_at_points) * capacities)

            # Update marker pop-ups with production values
            for marker, name, status, capacity, operator, owner, wind_speed, production, commission_date, id in zip(marker_cluster.markers, project_names, statuses, capacities, operators, owners, wind_speeds_at_points, productions, commission_dates, ids):
                marker.popup.value = \
                    f"<strong>Project Name:</strong> {name}<br>"\
                    f"<strong>Status:</strong> {status}<br>"\
                    f"<strong>Capacity:</strong> {capacity} MW<br>"\
                    f"<strong>Operator:</strong> {operator}<br>"\
                    f"<strong>Owner:</strong> {owner}<br>"\
                    f"<strong>Type:</strong> {turbine_type}<br>"\
                    f"<strong>Height:</strong> {hub_height} MW<br>"\
                    f"<strong>Comission Date:</strong> {commission_date}<br>"\
                    f"<strong>Diameter:</strong> {rotor_diameter}<br>"\
                    f"<strong>Wind speed forecast:</strong> {wind_speed}<br>"\
                    f"<strong>Production forecast:</strong> {production}<br>"\
                    f"<button onclick=\"Shiny.setInputValue(\'entire_forecast\', {{id: {id}}})\">Entire Forecast</button>"
                
            # Prepare heatmap data for "operating" wind plants
            heatmap_data = [(lat, lon, prod) for lat, lon, prod, status in zip(lats_plants, lons_plants, productions, statuses) if status == "operating"]

            # Remove existing heatmap layer if any and add new one
            for layer in m.layers:
                if isinstance(layer, Heatmap):
                    m.remove(layer)
            heatmap = Heatmap(locations=heatmap_data, radius=5, blur=5, max_zoom=10)
            m.add(heatmap)
        
        global initial
        # Initialise the map with the first time step
        if initial == 0:
            update_map(None)
            initial = 1
        
        # Observe slider value changes and update map
        slider.observe(update_map, names='value')

        # Add FullScreenControl to map
        m.add(FullScreenControl())

        return m
    

    ##### Page 2 #####

    forecast_data = reactive.Value({"wind_speeds": None, "productions": None})
    time_series = reactive.Value(None)
    button_status = reactive.Value("download")

    # Function to handle file upload
    @reactive.Effect
    @reactive.event(input.upload_file)
    def handle_file_upload():
        if input.upload_file() is not None:
            try:
                file = input.upload_file()[0]['datapath']
                # Attempt to read Excel file with specified columns
                time_series_data = pd.read_excel(file)
                h = time_series.get()
                print('vorher', h)
                time_series.set(time_series_data) # calls output_graph function
                h = time_series.get()
                print('nachher', h)
                ui.update_action_button("action_button", label="Contribute Data")
                button_status.set('contribute')
            except Exception as e:
                # Send feedback to UI if file format is incorrect
                ui.notification_show("Wrong file format, please download the example file for orientation.", duration=None)
    
    # Generate example file for download
    @reactive.Effect
    @reactive.event(input.download_example)
    async def generate_example_file():
        # Create example workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Example Time Series"
        
        # Add header and example data
        ws.append(["Date", "Production (kW)"])
        for date, production in zip(example_dates, example_production):
            ws.append([date, production])
        
        # Save workbook to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Encode and send for download
        file_data_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        await session.send_custom_message("download_file", {
            "data": file_data_base64,
            "filename": "example_time_series.xlsx"
        })

    # Observing slider changes to revert to forecast view
    @reactive.Effect
    @reactive.event(input.lat, input.lon, input.turbine_type, input.hub_height, input.commission_date, input.rotor_diameter, input.capacity)
    def observe_slider_changes():
        # reset reactive variables
        project_name.set(None)
        status.set(None)
        operator.set(None)
        owner.set(None)

        # display regular elements
        time_series.set(None) # calls output_graph function
        ui.update_action_button("action_button", label="Download Forecast")
        button_status.set('download')

    # Capture user input and display configuration summary
    @output
    @render.text
    def output_summary():
        # Capture inputs
        lat_plant = input.lat()
        lon_plant = input.lon()
        turbine_type = input.turbine_type()
        hub_height = input.hub_height()
        commission_date = input.commission_date()
        rotor_diameter = input.rotor_diameter()
        capacity = input.capacity()

        # Return configuration summary text
        summary_html = (
            f"<b>Turbine Configuration</b><br><br>"
            f"<b>Project Name:</b> {project_name.get()}<br>"
            f"<b>Status:</b> {status.get()}<br>"
            f"<b>Operator:</b> {operator.get()}<br>"
            f"<b>Owner:</b> {owner.get()}<br>"
            f"<b>Location:</b> ({lat_plant}, {lon_plant})<br>"
            f"<b>Type:</b> {turbine_type}<br>"
            f"<b>Hub Height:</b> {hub_height} m<br>"
            f"<b>Commission Date:</b> {commission_date}<br>"
            f"<b>Rotor Diameter:</b> {rotor_diameter} m<br>"
            f"<b>Capacity:</b> {capacity} kW<br>"
        )
        return ui.HTML(summary_html)

    # Plot forecasted production over time
    @output
    @render.plot
    def output_graph():
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.set_xlabel("Date")
        ax.set_ylabel("Production (kW)")

        # Add horizontal dashed line for capacity
        capacity = input.capacity()
        ax.axhline(y=capacity, color='gray', linestyle='--', label=f"Capacity ({capacity} kW)")

        time_series_data = time_series.get()

        if time_series_data is not None and not time_series_data.empty: # check if user has uploaded time series to plot it
            ax.plot(pd.to_datetime(time_series_data.iloc[:, 0], errors='coerce'), time_series_data.iloc[:, 1], label="Historical Production (kW)")
            print(pd.to_datetime(time_series_data.iloc[:, 0], errors='coerce'))
            print(pd.to_datetime(time_series_data.iloc[:, 0], errors='coerce').dtype)
            print(time_series_data.iloc[:, 1])
            print(time_series_data.iloc[:, 1].dtype)
            ax.set_title("Historical Wind Turbine Production")
        else: # Retrieve inputs
            lat_plant = input.lat()
            lon_plant = input.lon()

            # Calculate forecasted productions
            productions = []
            wind_speeds_at_points = []
            for step_index in range(len(valid_times_interpol)):
                spatial_interpolator = interp2d(lons, lats, total_selection_interpol[step_index], kind='cubic')
                wind_speeds_at_points.append(spatial_interpolator(lon_plant, lat_plant)[0])
                productions.append(np.abs(power_curve_normalised(wind_speeds_at_points[-1]) * capacity))

            # Store the forecast data in the reactive `forecast_data`
            forecast_data.set({"wind_speeds": wind_speeds_at_points, "productions": productions})
            ax.plot(valid_times_interpol, productions, label="Predicted Production (kW)")
            ax.set_title("Forecasted Wind Turbine Production")

        ax.legend()
        ax.grid(True)
        plt.tight_layout()
        return fig  # Returning the figure will embed it in the app

    def download_forecast_data():

        data = forecast_data.get()
        wind_speeds, productions = data["wind_speeds"], data["productions"]

        lat_plant = input.lat()
        lon_plant = input.lon()
        turbine_type = input.turbine_type()
        hub_height = input.hub_height()
        commission_date = input.commission_date()
        rotor_diameter = input.rotor_diameter()
        capacity = input.capacity()

        # Retrieve the actual values of reactive objects
        project_name_value = project_name.get()
        status_value = status.get()
        operator_value = operator.get()
        owner_value = owner.get()

        # Create a new workbook and add data
        wb = Workbook()
        
        # Sheet1: Turbine Specifications
        ws_specs = wb.active
        ws_specs.title = "Turbine Specifications"

        # Define the specs_data list with required values only
        specs_data = [
            ["Specification", "Value"],
            ["Project Name", project_name_value],
            ["Status", status_value],
            ["Operator", operator_value],
            ["Owner", owner_value],
            ["Location", f"({lat_plant}, {lon_plant})"],
            ["Type", turbine_type],
            ["Hub Height (m)", hub_height],
            ["Commission Date", commission_date],
            ["Rotor Diameter (m)", rotor_diameter],
            ["Capacity (kW)", capacity]
        ]

        # Populate the workbook
        for row in specs_data:
            ws_specs.append(row)

        # Sheet2: Production Forecast
        ws_forecast = wb.create_sheet("Production Forecast")

        # Add headers for date, wind speed, and production
        ws_forecast.append(["Date", "Wind Speed (m/s)", "Production (kW)"])

        # Add production data per time step
        for time, wind_speed, production in zip(valid_times_interpol, wind_speeds, productions):
            time_as_datetime = pd.to_datetime(time).to_pydatetime()
            ws_forecast.append([time_as_datetime, wind_speed, production])  # Ensure wind speed and production values are correctly added

        # Save in buffer and return
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # Function to trigger download on button click
    @reactive.Effect
    @reactive.event(input.action_button)
    async def action():
        if button_status.get() == "download":
        
            # Retrieve forecast data as bytes
            file_data = download_forecast_data()
            
            # Encode file in Base64
            file_data_base64 = base64.b64encode(file_data).decode('utf-8')
            
            # Send file as Base64 to client for download
            await session.send_custom_message("download_file", {
                "data": file_data_base64,
                "filename": "forecasted_production.xlsx"
            })
        
        else: # button_status.get() == "contribute"
            ui.notification_show(
                "Thank you for uploading your time series. It will be checked for plausibility and possibly used to improve the model at the next training.",
                duration=None
            )
            # save time_series somewhere until training
            time_series.set(None)

# Define absolute path to `www` folder
path_www = os.path.join(os.path.dirname(__file__), "www")

# Start the app with the `www` directory as static content
app = App(app_ui, server, static_assets={"/www": path_www})

if __name__ == "__main__":
    # Check if the variable `RENDER` is set to detect if the app is running on Render
    if os.getenv("RENDER") or os.getenv("WEBSITE_HOSTNAME"):  # for Render or Azure Server
        host = "0.0.0.0"  # For Render or other external deployments
    else:
        host = "127.0.0.1"  # For local development (localhost)

    app.run(host=host, port=8000)  # port binding: set the server port to 8000, because this is what Azure expects
