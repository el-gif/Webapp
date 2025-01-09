import io
import os
os.environ['FOR_DISABLE_CONSOLE_CTRL_HANDLER'] = 'T' # to disable Ctrl+C crashing python when having scipy.interpolate imported (disables Fortrun runtime library from intercepting Ctrl+C signal and lets it to the Python interpreter)
from openpyxl import Workbook
from shiny import ui, App, reactive, render
from shinywidgets import render_widget, output_widget
from ipyleaflet import Map, Marker, MarkerCluster, WidgetControl, FullScreenControl, AwesomeIcon, Heatmap
from ipywidgets import SelectionSlider, Play, VBox, jslink, Layout, HTML  # pip install ipywidgets==7.6.5, because version 8 has an issue with popups (https://stackoverflow.com/questions/75434737/shiny-for-python-using-add-layer-for-popus-from-ipyleaflet)
import numpy as np
import pandas as pd
import xarray as xr
from scipy.interpolate import interp2d  # pip install scipy==1.13.1, interp2d much faster than RegularGridInterpolator, even if deprecated
import base64
from ecmwf.opendata import Client
from branca.colormap import linear
import matplotlib.pyplot as plt
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
import time
import math
import joblib
import torch
import torch.nn as nn
from sklearn.preprocessing import OneHotEncoder

# Define initial variables
initial = 0
lat_min, lat_max = 35, 72
lon_min, lon_max = -25, 45

# Path to the stored weather data
wind_file = "data/weather_forecast/data_europe.grib2"
# Set time interval for updates (in seconds)
time_threshold = 24 * 3600  # 24 hours in seconds
overwrite = 0  # force overwriting of data

# Initialise ECMWF client
client = Client(
    source="ecmwf",
    model="ifs",
    resol="0p25"
)

if os.getenv("RENDER") or os.getenv("WEBSITE_HOSTNAME"):  # for Render or Azure Server
    # for saving storage on server, only 4 time steps
    step_selection = [0, 3, 6, 9]
else:
    # all available steps up to 6 days (144 hours)
    step_selection = [0, 3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33, 36, 39, 42, 45, 48, 51, 54, 57, 60,
                      63, 66, 69, 72, 75, 78, 81, 84, 87, 90, 93, 96, 99, 102, 105, 108, 111, 114, 117, 120,
                      123, 126, 129, 132, 135, 138, 141, 144]

# all available steps
# step_selection = [
#     0, 3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33, 36, 39, 42, 45, 48, 51, 54, 57, 60,
#     63, 66, 69, 72, 75, 78, 81, 84, 87, 90, 93, 96, 99, 102, 105, 108, 111, 114, 117, 120,
#     123, 126, 129, 132, 135, 138, 141, 144, 150, 156, 162, 168, 174, 180, 186, 192, 198, 204,
#     210, 216, 222, 228, 234, 240
# ]

# Function to check the age of the file, to avoid a cron-job, which is a paid feature on most Cloud Computing platforms
def is_data_stale(wind_file, time_threshold):
    if not os.path.exists(wind_file):
        # File does not exist, data must be fetched
        return True
    # Determine the age of the file (in seconds since the last modification)
    file_age = time.time() - os.path.getmtime(wind_file)
    # Check if the file is older than the specified threshold
    return file_age > time_threshold

if is_data_stale(wind_file, time_threshold) or overwrite:
    print("Data is outdated, does not exist, or should be overwritten. Fetching new data.")
    # Fetching API data since they are either missing or older than 10 hours
    result = client.retrieve(  # retrieve data worldwide, because no area tag available (https://github.com/ecmwf/ecmwf-opendata, https://www.ecmwf.int/en/forecasts/datasets/open-data)
        type="fc",
        param=["100v", "100u"],  # U- and V-components of wind speed
        target=wind_file,
        time=0,  # Forecast time (model run at 00z)
        step=step_selection
    )
    print("New data has been successfully fetched and saved.")
else:
    print("Data is current and will be used.")

# Load the wind data (Grib2 file)
ds = xr.open_dataset(wind_file, engine='cfgrib')

# Filter the data for Europe and extract relevant columns
ds_filtered = ds.sel(latitude=slice(lat_max, lat_min), longitude=slice(lon_min, lon_max))
lats = ds_filtered['latitude'].values
lons = ds_filtered['longitude'].values
u = ds_filtered['u100'].values
v = ds_filtered['v100'].values
valid_times = ds_filtered['valid_time'].values

# Filter the data for Europe and extract relevant columns
df = pd.read_parquet("data/WPPs/The_Wind_Power.parquet") # 0.7 seconds when WPPs already regionally filtered and stored as parquet file. As unfiltered excel file it takes 11 seconds
df = df[df["Status"] == "Production"]
df = df.iloc[::100] # only every 10th wpp is possible to alleviate computational and storage burden, not much more
ids = df['ID'].values
countries = df['Country'].values
project_names = df['Name'].values
lats_plants = df['Latitude'].values
lons_plants = df['Longitude'].values
manufacturers = df['Manufacturer'].values
turbine_types = df['Turbine'].values
hub_heights = df['Hub height'].values
numbers_of_turbines = df['Number of turbines']
capacities = df['Total power'].values / 1e3 # kW to MW as the model has been trained on and the capacity scaler has been fitted
developers = df['Developer'].values
operators = df['Operator'].values
owners = df['Owner'].values
commissioning_dates = df['Commissioning date'].values
ages_months = df['Ages months'].values
commissioning_date_statuses = df['Commissioning date status'].values
hub_height_statuses = df['Hub height status'].values
number_wpps = len(ids)

# Lade die gespeicherte Reihenfolge der Turbinentypen
known_turbine_types = np.load("model2/turbine_types_order.npy")
processed_turbine_types = [turbine if turbine in known_turbine_types else "nan" for turbine in turbine_types]
encoder = OneHotEncoder(categories=[known_turbine_types], sparse_output=False)
turbine_types_onehot = encoder.fit_transform(np.array(processed_turbine_types).reshape(-1, 1))

hub_height_min = math.floor(0.9 * df['Hub height'].min())
hub_height_max = math.ceil(1.1 * df['Hub height'].max())
commissioning_years = df['Commissioning date'].str.split('/').str[0].astype(int)
min_year = commissioning_years.min()
max_year = commissioning_years.max()
min_capacity = 0
max_capacity = capacities.max()

# Lade den Scaler
scalers = joblib.load("model2/scalers.pkl")

# Wende sie auf neue Daten an
scaled_ages_months = scalers["ages"].transform(ages_months.reshape(-1, 1)).flatten()
scaled_hub_heights = scalers["hub_heights"].transform(hub_heights.reshape(-1, 1)).flatten()

import torch.nn as nn

class MLP(nn.Module):
    def __init__(self, input_size, use_dropout=False, dropout_rate=0.3, 
                 use_batch_norm=False, activation_fn=nn.ReLU):
        super(MLP, self).__init__()

        layers = []

        # Erste Schicht
        layers.append(nn.Linear(input_size, 256))
        if use_batch_norm:
            layers.append(nn.BatchNorm1d(256))
        layers.append(activation_fn())

        # Zweite Schicht
        layers.append(nn.Linear(256, 128))
        if use_batch_norm:
            layers.append(nn.BatchNorm1d(128))
        layers.append(activation_fn())

        # Dritte Schicht
        layers.append(nn.Linear(128, 64))
        if use_batch_norm:
            layers.append(nn.BatchNorm1d(64))
        layers.append(activation_fn())

        # Dropout nach der letzten versteckten Schicht (optional)
        if use_dropout:
            layers.append(nn.Dropout(dropout_rate))

        # Ausgabeschicht
        layers.append(nn.Linear(64, 1))

        # Modell zusammenstellen
        self.model = nn.Sequential(*layers)

    def forward(self, x):
        return self.model(x)
    
# Lade die Metadaten
input_size = torch.load("model2/input_size", weights_only=True)
model = MLP(input_size=input_size)

# Lade die Modellgewichte
model.load_state_dict(torch.load("model2/trained_parameters.pth", weights_only=True))

model.eval()
print("Model loaded")

# Interpolation period and interval
start_time = valid_times[0]
end_time = valid_times[-1]
total_hours = int((end_time - start_time) / np.timedelta64(1, 'h'))
step_size = np.timedelta64(1, 'h')
step_size_hours = step_size / np.timedelta64(1, 'h')
valid_times_interpol = [start_time + i * step_size for i in range(int(total_hours / step_size_hours))]

# Calculate total wind speed and convert to 3D array
total_selection = np.array([np.sqrt(u_value**2 + v_value**2) for u_value, v_value in zip(u, v)])

# Function for temporal interpolation
def interpolation(list):
    list_interpol = []
    for i in range(len(list) - 1):
        step_start = valid_times[i]
        step_end = valid_times[i + 1]
        list_start = list[i]
        list_end = list[i + 1]
        list_interpol.append(list_start)
        interval = step_end - step_start
        for step in np.arange(step_start + step_size, step_end, step_size):
            factor = (step - step_start) / interval
            interpolated_value = (1 - factor) * list_start + factor * list_end
            list_interpol.append(interpolated_value)
    list_interpol.append(list[-1])
    return list_interpol

# Interpolation
total_selection_interpol = interpolation(total_selection)

valid_times_interpol = []
steps = int(total_hours / step_size_hours)
for i in range(steps):
    valid_times_interpol.append(start_time + i * step_size)

example_data = pd.read_parquet("data/production_history/Example/example_time_series.parquet")
example_dates = example_data['Date']
example_production = example_data['Production (kW)']

app_ui = ui.page_navbar(
    ui.nav_panel(
        "Spatial Forecast",
        output_widget("map")
    ),
    ui.nav_panel(
        "Temporal Forecast",
        ui.row(
            ui.column(2,  # Left column for input fields
                ui.input_numeric("lat", "Turbine Latitude", min=lat_min, max=lat_max, value=(lat_min + lat_max) / 2),
                ui.input_numeric("lon", "Turbine Longitude", min=lon_min, max=lon_max, value=(lon_min + lon_max) / 2),
                ui.input_select("turbine_type", "Turbine Type", choices=known_turbine_types.tolist(), selected=known_turbine_types[0]),
                ui.input_slider("hub_height", "Turbine Hub Height (m)", min=hub_height_min, max=hub_height_max, value=(hub_height_min + hub_height_max) / 2),
                ui.input_slider("commissioning_date_year", "Commissioning Date (Year)", min=min_year, max=max_year, value=(min_year + max_year) / 2, step=1),
                ui.input_slider("commissioning_date_month", "Commissioning Date (Month)", min=1, max=12, value=6, step=1),
                ui.input_slider("capacity", "Capacity (MW)", min=min_capacity, max=max_capacity, value=(min_capacity + max_capacity) / 2),
                ui.tags.br(),
                ui.input_file("upload_file", "Contribute data for this configuration", accept=[".xlsx"]),
                ui.input_action_link("download_example", "Download Example File")
            ),
            ui.column(10,  # Right column for output
                ui.panel_well(  # Panel to centre the content
                    ui.output_ui("output_summary"),
                    ui.tags.br(),
                    ui.output_plot("output_graph"),
                    ui.tags.br(),
                    ui.input_action_button("action_button", "Download Forecast")
                ),
            )
        ),
        value='customise_WPP'
    ),
    ui.nav_panel(
        "Documentation",
        ui.h3("Wind Power Forecast Application"),
        ui.HTML(
            "This web application visualises wind power production forecasts for wind power plants across Europe.<br>"
            "Users can view wind plant information, forecasted wind speeds, and production values on an interactive map.<br>"
            "Each wind plant can be customised by selecting its attributes, and predictions are displayed based on input parameters.<br>"
            "All given times are in Central European Time (UTC+1), please consider this especially, when contributing your own data.<br>"
            "The application is built using Python, Shiny, and Jupyter notebooks.<br>"
            "The model is trained on a dataset of wind power plants in Europe and uses weather forecast data from the European Centre for Medium-Range Weather Forecasts (ECMWF).<br>"
            "The model is trained on the following features: Turbine Type, Hub Height, Capacity, Age, and Wind Speed.<br>"
            "The model is a simple Multi-Layer Perceptron (MLP) with four hidden layers and ReLU activation functions.<br>"
            "Where the hub heigh is missing, the mean value of the wind power plants in the database is assumed to calculate the forecast.<br>"
            "Where the commissioning date is missing entirely, an estimated value of 2010/06 is assumed, where the month is missing, 06 is assumed.<br>"
            "Where the turbine type is new, nan is auto-selected in the dropdown menu to generate the forecast.<br>"
            "Please contribute to ameliorate the model, taking into account the time zone and hand in the production values in kW (see example file).<br>"
            "When setting extreme values for the input features, please consider that the model has barely been trained on such values and that the results will be less credible."
        )
    ),
    ui.head_content(
        ui.tags.script("""
            Shiny.addCustomMessageHandler("download_file", function(message) {
                var link = document.createElement('a');
                link.href = 'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,' + message.data;
                link.download = message.filename;
                document.body.appendChild(link);
                link.click();
                document.body.removeChild(link);
            });
        """),
        ui.tags.link(rel="icon", type="image/png", href="/www/WPP_icon2.png")
    ),
    id="navbar_selected",
    title="Wind Power Forecast"
)

# server function
def server(input, output, session):

    ##### Page 1 #####

    # Define reactive values for storing additional information
    project_name = reactive.Value(None)
    operator = reactive.Value(None)
    owner = reactive.Value(None)
    commissioning_date_status = reactive.Value(None)
    hub_height_status = reactive.Value(None)
    country = reactive.Value(None)

    is_programmatic_change = reactive.Value(False)

    @reactive.effect
    @reactive.event(input.entire_forecast)
    def entire_forecast_function():

        id = input.entire_forecast()['id']
        
        # Setzen des Flags: Änderungen sind programmatisch
        is_programmatic_change.set(True)

        index = list(ids).index(id)

        # Speichern der zusätzlichen Informationen in den reaktiven Werten
        project_name.set(project_names[index])
        operator.set(operators[index])
        owner.set(owners[index])
        commissioning_date_status.set(commissioning_date_statuses[index])
        hub_height_status.set(hub_height_statuses[index])
        country.set(countries[index])

        # Parameter extrahieren
        lat = lats_plants[index]
        lon = lons_plants[index]
        turbine_type = processed_turbine_types[index]
        hub_height = hub_heights[index]
        capacity = capacities[index]
        commissioning_date = commissioning_dates[index]
        commissioning_date_year, commissioning_date_month = commissioning_date.split("/")
        commissioning_date_year = int(commissioning_date_year)
        commissioning_date_month = int(commissioning_date_month)

        # Update der Eingabefelder mit neuen Werten vor Wechseln des Tabs
        ui.update_numeric("lat", value=lat)
        ui.update_numeric("lon", value=lon)
        ui.update_select("turbine_type", selected=turbine_type)
        ui.update_slider("hub_height", value=hub_height)
        ui.update_slider("commissioning_date_year", value=commissioning_date_year)
        ui.update_slider("commissioning_date_month", value=commissioning_date_month)
        ui.update_slider("capacity", value=capacity)

        # Wechsel zu "Customise WPP" Tab
        ui.update_navs("navbar_selected", selected="customise_WPP")

    @output
    @render_widget
    def map():
        m = Map(
            center=[(lat_min + lat_max) / 2, (lon_min + lon_max) / 2],
            zoom=5,
            layout=Layout(width='100%', height='95vh'),
            scroll_wheel_zoom=True
        )

        markers = []
        for name, capacity, operator, id, lat, lon in zip(project_names, capacities, operators, ids, lats_plants, lons_plants):
            
            popup_content = HTML(
                f"<strong>Project Name:</strong> {name}<br>"
                f"<strong>Capacity:</strong> {capacity} MW<br>"
                f"<strong>Operator:</strong> {operator}<br>"
                f"<strong>Wind speed forecast:</strong> select forecast step<br>"
                f"<strong>Production forecast:</strong> select forecast step<br>"
                f"<button onclick=\"Shiny.setInputValue('entire_forecast', {{id: {id}}})\">Entire Forecast</button>" # timestamp to always have a slightly different button value to ensure that each and every click on the "entire forecast" button triggers the event, even click on same button twice
            )

            marker = Marker(
                location=(lat, lon),
                popup=popup_content,
                rise_offset=True,
                draggable=False
            )
            markers.append(marker)

        # Cluster erstellen und zur Karte hinzufügen
        marker_cluster = MarkerCluster(markers=markers)
        m.add(marker_cluster)

        # Slider for time steps
        play = Play(min=0, max=total_hours, step=step_size_hours, value=0, interval=500, description='Time Step')
        slider = SelectionSlider(options=valid_times_interpol, value=valid_times_interpol[0], description='Time')
        jslink((play, 'value'), (slider, 'index'))
        slider_box = VBox([play, slider])
        m.add(WidgetControl(widget=slider_box, position='topright'))

        # Map update function based on slider
        def update_map(change):
            time_step = slider.value
            step_index = int((time_step - start_time) / step_size)
            spatial_interpolator = interp2d(lons, lats, total_selection_interpol[step_index], kind='cubic')
            wind_speeds_at_points = np.array([spatial_interpolator(lon, lat)[0] for lon, lat in zip(lons_plants, lats_plants)])
            scaled_wind_speeds_at_points = scalers["winds"].transform(wind_speeds_at_points.reshape(-1, 1)).flatten()

            all_input_features = np.hstack([
                turbine_types_onehot,
                scaled_hub_heights.reshape(-1, 1),
                scaled_ages_months.reshape(-1, 1),
                scaled_wind_speeds_at_points.reshape(-1, 1)
            ])

            input_tensor = torch.tensor(all_input_features, dtype=torch.float32)

            with torch.no_grad():
                predictions = np.minimum(model(input_tensor).numpy().flatten() * capacities, capacities)

            predictions[(predictions < 0) | (np.array(wind_speeds_at_points) > 25)] = 0
            
            # Update marker pop-ups with production values
            for marker, name, capacity, operator, wind_speed, prediction, id in zip(marker_cluster.markers, project_names, capacities, operators, wind_speeds_at_points, predictions, ids):
                marker.popup.value = \
                    f"<strong>Project Name:</strong> {name}<br>"\
                    f"<strong>Capacity:</strong> {capacity} MW<br>"\
                    f"<strong>Operator:</strong> {operator}<br>"\
                    f"<strong>Wind speed forecast:</strong> {wind_speed:.2f} m/s<br>"\
                    f"<strong>Production forecast:</strong> {prediction:.2f} MW<br>"\
                    f"<button onclick=\"Shiny.setInputValue('entire_forecast', {{id: {id}}})\">Entire Forecast</button>"
                
            # Prepare heatmap data for "operating" wind plants
            heatmap_data = [(lat, lon, prod) for lat, lon, prod in zip(lats_plants, lons_plants, predictions)]

            # Remove existing heatmap layer if any and add new one
            for layer in m.layers:
                if isinstance(layer, Heatmap):
                    m.remove(layer)
            heatmap = Heatmap(locations=heatmap_data, radius=5, blur=5, max_zoom=10)
            m.add(heatmap)
        
        global initial
        # Initialise the map with the first time step
        if initial == 0:
            update_map(None)
            initial = 1
        
        # Observe slider value changes and update map
        slider.observe(update_map, names='value')

        # Add FullScreenControl to map
        m.add(FullScreenControl())

        return m
    

    ##### Page 2 #####

    forecast_data = reactive.Value({"wind_speeds": None, "productions": None})
    time_series = reactive.Value(None)
    button_status = reactive.Value("download")

    # Function to handle file upload
    @reactive.Effect
    @reactive.event(input.upload_file)
    def handle_file_upload():
        if input.upload_file() is not None:
            try:
                file = input.upload_file()[0]['datapath']
                # Attempt to read Excel file with specified columns
                time_series_data = pd.read_excel(file)
                time_series.set(time_series_data) # calls output_graph function
                ui.update_action_button("action_button", label="Contribute Data")
                button_status.set('contribute')
            except Exception as e:
                # Send feedback to UI if file format is incorrect
                ui.notification_show("Wrong file / content format, please download the example file for orientation.", duration=None)
    
    # Generate example file for download
    @reactive.Effect
    @reactive.event(input.download_example)
    async def generate_example_file():
        # Create example workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Example Time Series"
        
        # Add header and example data
        ws.append(["Date", "Production (kW)"])
        for date, production in zip(example_dates, example_production):
            ws.append([date, production])
        
        # Save workbook to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Encode and send for download
        file_data_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        await session.send_custom_message("download_file", {
            "data": file_data_base64,
            "filename": "example_time_series.xlsx"
        })

    # Observing slider changes to revert to forecast view
    @reactive.Effect
    @reactive.event(input.lat, input.lon, input.turbine_type, input.hub_height, input.commissioning_date_year, input.commissioning_date_month, input.capacity)
    def observe_slider_changes():
        # Überspringen, wenn Änderungen programmatisch sind
        if is_programmatic_change.get():
            is_programmatic_change.set(False)
            return
        
        # reset reactive variables
        project_name.set(None)
        operator.set(None)
        owner.set(None)
        commissioning_date_status.set(None)
        hub_height_status.set(None)
        country.set(None)

        # display regular elements
        time_series.set(None) # calls output_graph function
        ui.update_action_button("action_button", label="Download Forecast")
        button_status.set('download')

        # Zurücksetzen des Flags
        is_programmatic_change.set(False)

    # Capture user input and display configuration summary
    @output
    @render.text
    def output_summary():
        # Capture inputs
        lat_plant = input.lat()
        lon_plant = input.lon()
        turbine_type = input.turbine_type()
        hub_height = input.hub_height()
        commissioning_date_year = input.commissioning_date_year()
        commissioning_date_month = input.commissioning_date_month()
        capacity = input.capacity()

        if hub_height_status.get() == 0:
            hub_height_display = f"{hub_height:.2f} m"
        else:
            hub_height_display = "nan"

        if commissioning_date_status.get() == 0:
            commissioning_date_display = f"{commissioning_date_year}/{commissioning_date_month}"
        elif commissioning_date_status.get() == 1:
            commissioning_date_display = str(commissioning_date_year)
        else:
            commissioning_date_display = "nan"

        # Return configuration summary text
        summary_html = (
            f"<b>Turbine Configuration</b><br><br>"
            f"<b>Project Name:</b> {project_name.get()}<br>"
            f"<b>Country:</b> {country.get()}<br>"
            f"<b>Operator:</b> {operator.get()}<br>"
            f"<b>Owner:</b> {owner.get()}<br>"
            f"<b>Location:</b> ({lat_plant}, {lon_plant})<br>"
            f"<b>Type:</b> {turbine_type}<br>"
            f"<b>Hub Height:</b> {hub_height_display}<br>"
            f"<b>Commissioning Date:</b> {commissioning_date_display}<br>"
            f"<b>Capacity:</b> {capacity} MW<br>"
        )
        return ui.HTML(summary_html)

    # Plot forecasted production over time
    @output
    @render.plot
    def output_graph():
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.set_xlabel("Date")
        ax.set_ylabel("Production (MW)")

        # Add horizontal dashed line for capacity
        capacity = input.capacity()
        ax.axhline(y=capacity, color='gray', linestyle='--', label=f"Capacity ({capacity} MW)")

        time_series_data = time_series.get()

        if time_series_data is not None and not time_series_data.empty: # check if user has uploaded time series to plot it
            ax.plot(pd.to_datetime(time_series_data.iloc[:, 0], errors='coerce'), time_series_data.iloc[:, 1] / 1e3, label="Historical Production (MW)")
            ax.set_title("Historical Wind Turbine Production")
        else: # Retrieve inputs
            lat_plant = input.lat()
            lon_plant = input.lon()
            turbine_type = input.turbine_type()
            hub_height = input.hub_height()
            commissioning_date_year = input.commissioning_date_year()
            commissioning_date_month = input.commissioning_date_month()
            ref_date = pd.Timestamp("2024-12-01")
            age_months = (ref_date.year - commissioning_date_year) * 12 + (ref_date.month - commissioning_date_month)
            capacity = input.capacity()

            scaled_hub_height = scalers["hub_heights"].transform(np.array([[hub_height]]))[0][0]
            scaled_age_months = scalers["ages"].transform(np.array([[age_months]]))[0][0]

            # Calculate forecasted productions
            wind_speeds_at_point = []
            num_steps = len(valid_times_interpol)
            for step_index in range(num_steps):
                spatial_interpolator = interp2d(lons, lats, total_selection_interpol[step_index], kind='cubic')
                wind_speeds_at_point.append(spatial_interpolator(lon_plant, lat_plant)[0])

            scaled_wind_speeds_at_point = scalers["winds"].transform(np.array(wind_speeds_at_point).reshape(-1, 1)).flatten()

            turbine_type_repeated = np.tile(encoder.transform(np.array([[turbine_type]])), (num_steps, 1))
            hub_height_repeated = np.full((num_steps, 1), scaled_hub_height)
            age_repeated = np.full((num_steps, 1), scaled_age_months)
            wind_speed_column = scaled_wind_speeds_at_point.reshape(-1, 1)

            all_input_features = np.hstack([
                turbine_type_repeated,
                hub_height_repeated,
                age_repeated,
                wind_speed_column
            ])

            np.save("all_input_features.npy", all_input_features)

            input_tensor = torch.tensor(all_input_features, dtype=torch.float32)

            with torch.no_grad():
                predictions = np.minimum(model(input_tensor).numpy().flatten() * capacity, np.full(num_steps, capacity))

            predictions[(predictions < 0) | (np.array(wind_speeds_at_point) > 25)] = 0

            # Store the forecast data in the reactive `forecast_data`
            forecast_data.set({"wind_speeds": wind_speeds_at_point, "productions": predictions})
            ax.plot(valid_times_interpol, predictions, label="Predicted Production (MW)")
            ax.set_ylim(bottom=0)
            ax.set_title("Forecasted Wind Turbine Production")

        ax.legend()
        ax.grid(True)
        plt.tight_layout()
        return fig  # Returning the figure will embed it in the app

    def download_forecast_data():

        data = forecast_data.get()
        wind_speeds, productions = data["wind_speeds"], data["productions"]

        lat_plant = input.lat()
        lon_plant = input.lon()
        turbine_type = input.turbine_type()
        hub_height = input.hub_height()
        commissioning_date_year = input.commissioning_date_year()
        commissioning_date_month = input.commissioning_date_month()
        capacity = input.capacity()

        # Create a new workbook and add data
        wb = Workbook()
        
        # Sheet1: Turbine Specifications
        ws_specs = wb.active
        ws_specs.title = "Turbine Specifications"

        # Define the specs_data list with required values only
        specs_data = [
            ["Specification", "Value"],
            ["Project Name", project_name.get()],
            ["Operator", operator.get()],
            ["Owner", owner.get()],
            ["Location", f"({lat_plant}, {lon_plant})"],
            ["Type", turbine_type],
            ["Hub Height (m)", round(hub_height, 2) if hub_height_status.get() == 0 else "nan"],
            ["Commissioning Date", f"{commissioning_date_year}/{commissioning_date_month}" if commissioning_date_status.get() == 0 else (str(commissioning_date_year) if commissioning_date_status.get() == 1 else "nan")],
            ["Capacity (MW)", capacity]
        ]

        # Populate the workbook
        for row in specs_data:
            ws_specs.append(row)

        # Sheet2: Production Forecast
        ws_forecast = wb.create_sheet("Production Forecast")

        # Add headers for date, wind speed, and production
        ws_forecast.append(["Date", "Wind Speed (m/s)", "Production (MW)"])

        # Add production data per time step
        for time, wind_speed, production in zip(valid_times_interpol, wind_speeds, productions):
            time_as_datetime = pd.to_datetime(time).to_pydatetime()
            ws_forecast.append([time_as_datetime, wind_speed, production])  # Ensure wind speed and production values are correctly added

        # Save in buffer and return
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # Function to trigger download on button click
    @reactive.Effect
    @reactive.event(input.action_button)
    async def action():
        if button_status.get() == "download":
        
            # Retrieve forecast data as bytes
            file_data = download_forecast_data()
            
            # Encode file in Base64
            file_data_base64 = base64.b64encode(file_data).decode('utf-8')
            
            # Send file as Base64 to client for download
            await session.send_custom_message("download_file", {
                "data": file_data_base64,
                "filename": "forecasted_production.xlsx"
            })
        
        else: # button_status.get() == "contribute"
            ui.notification_show(
                "Thank you for uploading your time series. It will be checked for plausibility and possibly used to improve the model at the next training.",
                duration=None
            )
            # save time_series somewhere until training
            time_series.set(None)

# Define absolute path to `www` folder
path_www = os.path.join(os.path.dirname(__file__), "www")

# Start the app with the `www` directory as static content
app = App(app_ui, server, static_assets={"/www": path_www})

if __name__ == "__main__":
    # Check if the variable `RENDER` is set to detect if the app is running on Render
    if os.getenv("RENDER") or os.getenv("WEBSITE_HOSTNAME"):  # for Render or Azure Server
        host = "0.0.0.0"  # For Render or other external deployments
    else:
        host = "127.0.0.1"  # For local development (localhost)

    app.run(host=host, port=8000)  # port binding: set the server port to 8000, because this is what Azure expects
